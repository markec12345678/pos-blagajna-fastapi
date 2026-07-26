from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import PlainTextResponse, Response
from sqlalchemy.orm import Session
from app.core.database import get_db
from app.models.order import Order, OrderItem
from app.models.inventory import Ingredient
from app.models.waste import WasteRecord
from app.models.shift import EmployeeShift
from app.models.customer import Customer
from app.models.loyalty import LoyaltyTransaction
from app.models.reservation import Reservation
from app.models.table_model import TableModel
from app.models.user import User
from datetime import datetime, timedelta
import csv, io

router = APIRouter(prefix="/export", tags=["export"])


def _get_sales_data(db, since, branch_id):
    q = db.query(Order).filter(Order.created_at >= since)
    if branch_id:
        q = q.filter(Order.branch_id == branch_id)
    return q.order_by(Order.created_at.desc()).all()


def _get_inventory_data(db, branch_id):
    q = db.query(Ingredient)
    if branch_id:
        q = q.filter(Ingredient.branch_id == branch_id)
    return q.order_by(Ingredient.name).all()


def _get_waste_data(db, since):
    return db.query(WasteRecord).filter(WasteRecord.created_at >= since).order_by(WasteRecord.created_at.desc()).all()


def _get_labor_data(db, since, branch_id):
    q = db.query(EmployeeShift).filter(EmployeeShift.clock_in >= since, EmployeeShift.clock_out != None)
    if branch_id:
        bu = {u.id for u in db.query(User).filter(User.branch_id == branch_id).all()}
        q = q.filter(EmployeeShift.user_id.in_(bu))
    return q.order_by(EmployeeShift.clock_in.desc()).all()


def _get_customer_data(db, branch_id):
    q = db.query(Customer)
    return q.order_by(Customer.name).all()


def _get_loyalty_data(db, since):
    return db.query(LoyaltyTransaction).filter(LoyaltyTransaction.created_at >= since).order_by(LoyaltyTransaction.created_at.desc()).all()


def _get_reservation_data(db, since, branch_id):
    q = db.query(Reservation).filter(Reservation.created_at >= since)
    if branch_id:
        q = q.filter(Reservation.branch_id == branch_id)
    return q.order_by(Reservation.reservation_time).all()


@router.get("/{export_type}")
def export_csv(export_type: str, days: int = 30, branch_id: int = 0, format: str = "csv", db: Session = Depends(get_db)):
    since = datetime.now() - timedelta(days=days)

    if format == "xlsx":
        return _export_xlsx(export_type, days, branch_id, since, db)
    return _export_csv(export_type, days, branch_id, since, db)


def _export_csv(export_type, days, branch_id, since, db):
    buf = io.StringIO()
    w = csv.writer(buf)

    if export_type == "sales":
        w.writerow(["ID", "Datum", "Tip", "Miza", "Stranka", "Skupaj", "Status", "Postavke"])
        for o in _get_sales_data(db, since, branch_id):
            items = ", ".join(f"{oi.item_name}x{oi.quantity}" for oi in (o.items or []))
            w.writerow([o.id, str(o.created_at), o.order_type, o.table_id or "", o.customer_name or "", o.total or 0, o.status, items])

    elif export_type == "inventory":
        w.writerow(["ID", "Sestavina", "Kategorija", "Zaloga", "Min. zaloga", "Enota", "Cena/enoto", "Vrednost"])
        for i in _get_inventory_data(db, branch_id):
            w.writerow([i.id, i.name, i.category, i.stock, i.min_stock, i.unit, i.cost_per_unit, round(i.stock * (i.cost_per_unit or 0), 2)])

    elif export_type == "waste":
        w.writerow(["ID", "Sestavina", "Kolicina", "Strosek", "Vzrok", "Opombe", "Datum"])
        for r in _get_waste_data(db, since):
            ing = db.query(Ingredient).filter(Ingredient.id == r.ingredient_id).first()
            w.writerow([r.id, ing.name if ing else "?", r.quantity, r.cost, r.reason, r.notes or "", str(r.created_at)])

    elif export_type == "labor":
        w.writerow(["ID", "Uporabnik", "Vloga", "Prihod", "Odhod", "Ure", "Urna postavka", "Strosek"])
        for s in _get_labor_data(db, since, branch_id):
            u = db.query(User).filter(User.id == s.user_id).first()
            hrs = round((s.clock_out - s.clock_in).total_seconds() / 3600, 2) if s.clock_in and s.clock_out else 0
            rate = u.hourly_rate or 10 if u else 10
            w.writerow([s.id, u.full_name if u else f"#{s.user_id}", u.role if u else "",
                        str(s.clock_in), str(s.clock_out), hrs, rate, round(hrs * rate, 2)])

    elif export_type == "customers":
        w.writerow(["ID", "Ime", "Telefon", "Email", "Naslov", "Clan", "Tocke", "Skupaj porabljeno", "Datum registracije"])
        for c in _get_customer_data(db, branch_id):
            w.writerow([c.id, c.name, c.phone or "", c.email or "", c.address or "",
                        "Da" if c.is_member else "Ne", c.loyalty_points or 0, c.total_spent or 0,
                        str(c.created_at) if c.created_at else ""])

    elif export_type == "loyalty":
        w.writerow(["ID", "Stranka ID", "Ime stranke", "Tocke", "Tip", "Narocilo ID", "Opomba", "Datum"])
        for t in _get_loyalty_data(db, since):
            c = db.query(Customer).filter(Customer.id == t.customer_id).first()
            w.writerow([t.id, t.customer_id, c.name if c else f"#{t.customer_id}",
                        t.points, t.type, t.order_id or "", t.note or "",
                        str(t.created_at) if t.created_at else ""])

    elif export_type == "reservations":
        w.writerow(["ID", "Miza", "Ime stranke", "Telefon", "Gostje", "Cas", "Status", "Opombe", "Datum"])
        for r in _get_reservation_data(db, since, branch_id):
            t = db.query(TableModel).filter(TableModel.id == r.table_id).first() if r.table_id else None
            w.writerow([r.id, t.name if t else "-", r.customer_name, r.customer_phone or "",
                        r.guests, str(r.reservation_time), r.status, r.notes or "",
                        str(r.created_at) if r.created_at else ""])

    else:
        raise HTTPException(400, "Invalid export type. Choose: sales, inventory, waste, labor, customers, loyalty, reservations")

    return PlainTextResponse(buf.getvalue(), media_type="text/csv",
                             headers={"Content-Disposition": f"attachment; filename={export_type}_{days}d.csv"})


def _export_xlsx(export_type, days, branch_id, since, db):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = export_type.capitalize()

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="059669", end_color="059669", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin", color="E2E8F0"),
        right=Side(style="thin", color="E2E8F0"),
        top=Side(style="thin", color="E2E8F0"),
        bottom=Side(style="thin", color="E2E8F0"),
    )

    if export_type == "sales":
        headers = ["ID", "Datum", "Tip", "Miza", "Stranka", "Skupaj", "Status", "Postavke"]
        ws.append(headers)
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border
        for o in _get_sales_data(db, since, branch_id):
            items = ", ".join(f"{oi.item_name}x{oi.quantity}" for oi in (o.items or []))
            ws.append([o.id, str(o.created_at), o.order_type, o.table_id or "", o.customer_name or "", o.total or 0, o.status, items])
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=6):
            row[5].number_format = '#,##0.00 €'

    elif export_type == "inventory":
        headers = ["ID", "Sestavina", "Kategorija", "Zaloga", "Min. zaloga", "Enota", "Cena/enoto", "Vrednost"]
        ws.append(headers)
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border
        for i in _get_inventory_data(db, branch_id):
            ws.append([i.id, i.name, i.category, i.stock, i.min_stock, i.unit, i.cost_per_unit, round(i.stock * (i.cost_per_unit or 0), 2)])
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=8):
            row[3].number_format = '#,##0.00'
            row[6].number_format = '#,##0.00 €'
            row[7].number_format = '#,##0.00 €'
            if row[3].value is not None and row[4].value is not None and row[3].value < row[4].value:
                row[3].font = Font(color="EF4444", bold=True)

    elif export_type == "waste":
        headers = ["ID", "Sestavina", "Kolicina", "Strosek", "Vzrok", "Opombe", "Datum"]
        ws.append(headers)
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border
        for r in _get_waste_data(db, since):
            ing = db.query(Ingredient).filter(Ingredient.id == r.ingredient_id).first()
            ws.append([r.id, ing.name if ing else "?", r.quantity, r.cost, r.reason, r.notes or "", str(r.created_at)])
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=4):
            row[3].number_format = '#,##0.00 €'

    elif export_type == "labor":
        headers = ["ID", "Uporabnik", "Vloga", "Prihod", "Odhod", "Ure", "Urna postavka", "Strosek"]
        ws.append(headers)
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border
        for s in _get_labor_data(db, since, branch_id):
            u = db.query(User).filter(User.id == s.user_id).first()
            hrs = round((s.clock_out - s.clock_in).total_seconds() / 3600, 2) if s.clock_in and s.clock_out else 0
            rate = u.hourly_rate or 10 if u else 10
            ws.append([s.id, u.full_name if u else f"#{s.user_id}", u.role if u else "",
                        str(s.clock_in), str(s.clock_out), hrs, rate, round(hrs * rate, 2)])
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=8):
            row[5].number_format = '#,##0.00'
            row[6].number_format = '#,##0.00 €'
            row[7].number_format = '#,##0.00 €'

    elif export_type == "customers":
        headers = ["ID", "Ime", "Telefon", "Email", "Naslov", "Clan", "Tocke", "Skupaj porabljeno", "Datum"]
        ws.append(headers)
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border
        for c in _get_customer_data(db, branch_id):
            ws.append([c.id, c.name, c.phone or "", c.email or "", c.address or "",
                       "Da" if c.is_member else "Ne", c.loyalty_points or 0, c.total_spent or 0,
                       str(c.created_at) if c.created_at else ""])
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=8):
            row[7].number_format = '#,##0.00 €'

    elif export_type == "loyalty":
        headers = ["ID", "Stranka ID", "Ime stranke", "Tocke", "Tip", "Narocilo ID", "Opomba", "Datum"]
        ws.append(headers)
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border
        for t in _get_loyalty_data(db, since):
            c = db.query(Customer).filter(Customer.id == t.customer_id).first()
            ws.append([t.id, t.customer_id, c.name if c else f"#{t.customer_id}",
                       t.points, t.type, t.order_id or "", t.note or "",
                       str(t.created_at) if t.created_at else ""])

    elif export_type == "reservations":
        headers = ["ID", "Miza", "Ime stranke", "Telefon", "Gostje", "Cas", "Status", "Opombe", "Datum"]
        ws.append(headers)
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border
        for r in _get_reservation_data(db, since, branch_id):
            t = db.query(TableModel).filter(TableModel.id == r.table_id).first() if r.table_id else None
            ws.append([r.id, t.name if t else "-", r.customer_name, r.customer_phone or "",
                       r.guests, str(r.reservation_time), r.status, r.notes or "",
                       str(r.created_at) if r.created_at else ""])

    else:
        raise HTTPException(400, "Invalid export type")

    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={export_type}_{days}d.xlsx"}
    )
